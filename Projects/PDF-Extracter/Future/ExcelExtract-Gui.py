import openpyxl
import tkinter as tk
from tkinter import filedialog

# Open a file dialog to select the .xlsx file
root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

# Load the workbook
wb = openpyxl.load_workbook(file_path)

# Select the sheet
sheet = wb["Sheet1"]

# Get the values of Cell 1 and Cell 2
ActualSales = sheet.cell(row=27, column=1).value
BudgetSales = sheet.cell(row=30, column=1).value

if ActualSales is None or not isinstance(ActualSales, (int, float)):
    raise TypeError("Cell 1 is not numeric")

if BudgetSales is None or not isinstance(BudgetSales, (int, float)):
    raise TypeError("Cell 2 is not numeric")

# Compare the values of Cell 1 and Cell 2
if BudgetSales < ActualSales:
    print(f"This Region HIT sales: Budgeted Sales ({BudgetSales}) Actual Sales({ActualSales})")
elif cell2 > cell1:
    print(f"This Region MISSED sales. Budgeted Sales: ({BudgetSales}) vs Actual Sales: ({ActualSales})")
else:
    print(f"This Region HIT sales. ({BudgetSales}) is equal to Cell 1 ({ActualSales})")
