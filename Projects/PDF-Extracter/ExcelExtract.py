import os
import openpyxl

# Get the list of all .xlsx files in the current directory
files = [f for f in os.listdir() if f.endswith(".xlsx")]

# Display the list of files
for i, file in enumerate(files):
    print(f"{i+1}. {file}")

# Ask the user to choose a file
file_index = int(input("Enter the number of the file you want to open: "))
file_path = files[file_index - 1]

# Load the workbook
wb = openpyxl.load_workbook(file_path)

# Select the sheet
sheet = wb["Sheet1"]

# Get the values of Cell 1 and Cell 2
region = sheet.cell(row=13, column=2).value

ActualSales = sheet.cell(row=27, column=1).value
BudgetSales = sheet.cell(row=30, column=1).value

ActualMTO = sheet.cell(row=575, column=1).value
BudgetMTO = sheet.cell(row=665, column=1).value

ActualOTHERFRESH = sheet.cell(row=577, column=1).value
BudgetOTHERFRESH = sheet.cell(row=667, column=1).value

ActualShrink = sheet.cell(row=1535, column=1).value
BudgetShrink = sheet.cell(row=1545, column=1).value

ActualPerson = sheet.cell(row=190, column=1).value
BudgetPerson = sheet.cell(row=216, column=1).value

# Add up the values of MTO and Other Fresh
actual_freshtotal = ActualMTO + ActualOTHERFRESH
budget_freshtotal = BudgetMTO + BudgetOTHERFRESH

# if ActualSales is None or not isinstance(ActualSales, (int, float)):
#     raise TypeError("Cell 1 is not numeric")

# if BudgetSales is None or not isinstance(BudgetSales, (int, float)):
#     raise TypeError("Cell 2 is not numeric")

# if ActualSales2 is None or not isinstance(ActualSales2, (int, float)):
#     raise TypeError("Cell 3 is not numeric")

# if BudgetSales2 is None or not isinstance(BudgetSales2, (int, float)):
#     raise TypeError("Cell 4 is not numeric")

if BudgetSales < ActualSales:
    print(f"Region ({region}) HIT Inside Sales. Budgeted Total: ({BudgetSales}) Actual Total: ({ActualSales})")
elif BudgetSales > ActualSales:
    print(f"Region ({region}) MISSED Inside Sales. Budgeted Total: ({BudgetSales}) Actual Total: ({ActualSales})")
else:
    print(f"Region ({region}) HIT Inside Sales. Budgeted Total ({BudgetSales}) Actual Total({ActualSales})")

if budget_freshtotal < actual_freshtotal:
    print(f"Region ({region}) HIT Total Fresh Sales. Budgeted Total: ({budget_freshtotal}) Actual Total: ({actual_freshtotal})")
elif budget_freshtotal > actual_freshtotal:
    print(f"Region ({region}) MISSED Total Fresh Sales. Budgeted Total: ({budget_freshtotal}) Actual Total: ({actual_freshtotal})")
else:
    print(f"Region ({region}) HIT Total Fresh Sales.  Budgeted Total: ({budget_freshtotal}) Actual Total: ({actual_freshtotal})")

#reverse the less/greater than for these

if BudgetShrink < ActualShrink:
    print(f"Region ({region}) HIT Shink %. Budgeted: ({BudgetShrink}) Actual: ({ActualShrink})")
elif BudgetShrink > ActualShrink:
    print(f"Region ({region}) MISSED Shrink %. Budgeted: ({BudgetShrink}) Actual: ({ActualShrink})")
else:
    print(f"Region ({region}) HIT Shrink %. Budgeted: ({BudgetShrink}) Actual: ({ActualShrink})")

if BudgetPerson > ActualPerson:
    print(f"Region ({region}) HIT Personnel %: Budgeted %: ({BudgetPerson}) Actual %: ({ActualPerson})")
elif BudgetPerson < ActualPerson:
    print(f"Region ({region}) MISSED Personnel %. Budgeted %: ({BudgetPerson}) Actual %: ({ActualPerson})")
else:
    print(f"Region ({region}) HIT Personnel %. Budgeted %: ({BudgetPerson}) Actual %: ({ActualPerson})")
